from openpyxl import *
from openpyxl.utils import range_boundaries
from pathlib import Path
import selenium_fcs as sf

class Direct:

    course = ''
    path = ''
    wb = ''
    sheet = ''
    list_insts = ''         # список институтов
    inst = ''               # выбранный институт
    list_edup = ''
    edup = ''
    list_napr = ''
    napr = ''
    list_groups = ''
    group = ''

    def set_path(self, path):
        self.path = path

    def set_course(self, course):
        self.course = course

    def set_inst(self, inst):
        self.inst = inst
        self.sheet = self.wb[inst]

    def set_napr(self, napr):
        self.napr = napr

    def set_edup(self, edup):
        self.edup = edup

    def set_group(self, group):
        self.group = group

    def get_list_inst(self):
        full_inst_list = self.wb.sheetnames
        # Иногда попадаются файлы с лишними страницами, тут их удаляем
        for s in full_inst_list:
            if s.startswith('3') or s == 'None': # если строка начинается с "3". в файле с расписанием 3 курса строки не начинаются с 3
                full_inst_list.remove(s)
        self.list_insts = full_inst_list

    def get_list_edup(self):
        y = (self.get_indexes(self.sheet, 'Образовательная программа'))[0]
        x = (self.get_indexes(self.sheet, self.napr))[1]
        gx = (self.next_idx(self.sheet, self.napr))[1]
        dict_of = {}
        temp = ''
        for j in range(x, gx):                           # сдвигаемся вправо на два элемента(ЭТО НАДО ИСПРАВИТЬ)

              name = str(self.sheet.cell(y,j).value).strip()

              if (name != '' and name != temp and name!='None'):
                   dict_of[name] = [y, j]
              else:
                   continue
              temp = name                                                 #сохраняем значение для проверки на объединенную ячейку

        self.list_edup = list(dict_of.keys())

    # Получение словаря {'название инст/напр': его индекс относительно талицы(строка, столбец)}
    def get_list_napr(self):
        # 1. определяем строку с направлениями
        # 2. определяем столбец с которого начинаются наименования
        y = (self.get_indexes(self.sheet, 'НАПРАВЛЕНИЕ'))[0]                  #индекс строки
        x = (self.next_idx(self.sheet, 'НАПРАВЛЕНИЕ'))[1]                     #индекс столбца
        dict_of = {}
        temp = ''
        for j in range(x, self.sheet.max_column+1):                           # сдвигаемся вправо на два элемента(ЭТО НАДО ИСПРАВИТЬ)

              name = str(self.sheet.cell(y,j).value).strip()

              if (name != '' and name != temp and name!='None'):
                   dict_of[name] = [y, j]
              else:
                   continue
              temp = name                                                 #сохраняем значение для проверки на объединенную ячейку
        self.list_napr = list(dict_of.keys())

    def get_list_group(self):
        # 1. получить строку на которой хранятся группы
        # 2. получить начало(индекс столбца) выбранного направления
        # 3. получить конец(индекс столбца) выбранного направления
        g_idx = (self.get_indexes(self.sheet, '№ учебной группы'))[0]    # строка
        s_idx = (self.get_indexes(self.sheet, self.edup))[1]             # начало(столбец)
        e_idx = (self.next_idx(self.sheet, self.edup))[1]                # конец(столбец)
        groups = []
        for i in range(s_idx, e_idx):
            group = self.sheet.cell(row = g_idx, column = i).value
            if group == "None": continue
            groups.append(group)
        self.list_groups = groups

    def check_name_in_list(self, name, somelist):
        for i in somelist:
            if i.find(name) != -1:
                return True
        return False

    #  БЛОК ПОЛУЧЕНИЯ И ОБРАБОТКИ ФАЙЛА ==========================================================
    def first_start(self):

        self.path = self.get_file_path()             # записываем путь скачанного файла
        self.unmerge_all_cells()                     # обрабатываем объединенные ячейки
        self.unmerge_institutes()                    # разделяем два института, хранящихся на одном листе
        print(f'path: {self.path}')
        self.wb = load_workbook(self.path)

    def clean_all(self):
        self.path = ''
        self.wb = ''
        self.sheet = ''
        self.list_insts = ''  # список институтов
        self.inst = ''  # выбранный институт
        self.list_edup = ''
        self.edup = ''
        self.list_napr = ''
        self.napr = ''
        self.list_groups = ''
        self.group = ''

    # Получение пути интересующего нас файла
    def get_file_path(self):
        with Path(r"/Users/katherine.marakhova/PycharmProjects/exampleBot/files") as direction:
            s = str(self.course) + "-курс-бакалавриат*.xlsx"
            for f in direction.glob(s):
                return f
            sf.get_file(self.course)  # скачиваем файл с сайта относительно курса
            for f in direction.glob(s):
                return f

    # Обработка объединенных ячеек, создание
    def unmerge_all_cells(self):
        workbook = load_workbook(self.path)
        n = len(workbook.sheetnames)
        for i in range(n):
            sheet = workbook.worksheets[i]
            set_merges = sorted(sheet.merged_cells.ranges.copy())
            for cell_group in set_merges:
                min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                sheet.unmerge_cells(str(cell_group))
                for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value
            workbook.save(self.path)

    # Разделение институтов, хранящихся на одном листе
    def unmerge_institutes(self):
        workbook = load_workbook(self.path)
        for s in workbook.sheetnames:
            if s.find(',') != -1:  # название нашего листа 'ИУПСиБК, ИИС 4 курс'
                sheet1 = workbook[s]

                t = s.find('курс') - 2  # вычленяем курс из названия листа
                z = s.find(',')  # находим позицию запятой
                fname = s[:z] + ' ' + s[t:]  # иупсибк
                sname = s[z + 2:]  # иис
                # Теперь лист на котором мы были будет называться как sname, а новый как fname
                sheet1.title = sname  # переименовыем в ИИС 4 курс
                workbook.create_sheet(fname)  # создаем лист ИУПСиБК 4 курс
                workbook.save(self.path)  # страхуемся, сохраняем наш док

                # переопределяем листы, так четче видно что где
                sheet1 = workbook[sname]  # иис тут заполнено
                sheet2 = workbook[fname]  # иупсибк тут пусто

                last_inst_name = ''
                y, x = self.get_indexes(sheet1, 'ИНСТИТУТ')
                for i in range(1, sheet1.max_column):
                    # val = sheet1[y][i].value
                    val = sheet1.cell(y, i).value
                    if val != 'None':
                        last_inst_name = val

                inst_idx = self.get_indexes(sheet1, last_inst_name)  # индекс с которого начинается второй институт(ИИС)

                # print(f'Последний институт {last_inst_name}, находится {inst_idx}')

                # заполняем новый лист(ИУПСиБК)
                for i in range(1, sheet1.max_row):
                    for j in range(1, inst_idx[1]):
                        sheet2.cell(i, j).value = sheet1.cell(i, j).value

                print(f'inst_idx[1]: {inst_idx[1]}')
                # удлаляем ненужные столбцы с исходного листа
                sheet1.delete_cols(idx=5, amount=(inst_idx[1] - 5))  # тут пока костыль в виде 4 - именно столько столбцов нужно отступить слева
                # надо будет написать функцию добывающую этот индекс, чтобы было гибко
                workbook.save(self.path)

    # ✔️️ Получение индекса(строка, столбец) относительно содержимого ячейки
    def get_indexes(self, sheet, header_el):
        header_el = (header_el.strip()).lower()
        for i in range(1, sheet.max_row):
            for j in range(1, sheet.max_column):
                val = (str(sheet.cell(i, j).value).strip()).lower()
                if (val == header_el):
                    return (i, j)

    # делает то же самое что и get_indexes только с понижением строки
    def get_indexes_cat(self, sheet, header_el, cat = 'Образовательная программа'):
        header_el = (header_el.strip()).lower()
        y, x = self.get_indexes(self.sheet, cat)    # нам нужна только строка! те у
        for i in range(y, sheet.max_row):
            for j in range(1, sheet.max_column):
                val = str(sheet.cell(i, j).value).strip().lower()
                # if (val == header_el or val.find(header_el) != -1):
                if (val == header_el):
                    return (i, j)

    # Получение индекса другого элемента(не равного по названию)
    def next_idx(self, sheet, name_el):
        name_el = name_el.strip().lower()
        y, x = self.get_indexes(sheet, name_el)  # строка, столбец начала
        for j in range(x, sheet.max_column):  # строка не меняется
            val = str(sheet.cell(row=y, column=j).value).strip().lower()
            if val != name_el and val != 'None' and val.find(name_el) == -1:
                return (y, j)
        return (y, sheet.max_column+1)  # значит он последний

        # Получение индекса другого элемента(не равного по названию)
    def next_idx_cat(self, sheet, name_el, category):
        row = self.get_indexes(self.sheet, category)[0]
        name_el = name_el.strip().lower()
        coll = self.get_indexes(sheet, name_el)[1]  # строка, столбец начала

        for c in range(coll, sheet.max_column+1):  # строка не меняется
            val = str(sheet.cell(row, c).value).strip().lower()
            if val == name_el: continue
            elif val != name_el: return (row, c)
        # if str(sheet.cell(row, c).value).strip().lower() == 'none': return (row, c-3)

    # БЛОК ВЫВОДА РАСПИСАНИЯ ======================================================================
    def get_scd_full(self):
        answer = f'Расписание для {self.edup.title()} {self.course}-{self.group}\n'
        # получаем граничные индексы
        c = self.get_indexes_cat(self.sheet, self.edup)[1]  # coll
        r = self.get_indexes(self.sheet, 'Понедельник')[0]  # row
        c = c + int(self.group) - 1  # перезаписываем столбец на новое значение

        temp = ''
        lesson = ''
        for i in range(r, self.sheet.max_row):

            day = str(self.sheet.cell(i, 2).value)
            time = str(self.sheet.cell(i, 3).value)
            even = str(self.sheet.cell(i, 4).value)
            sbj = str(self.sheet.cell(i, c).value)

            if (day == 'None'): break

            if sbj == 'None': sbj = 'Занятий нет'

            dash = '--------------------------------------️'

            if temp != day:
                spaces = ''
                n = len(dash) - len('❗️' + day + '❗️') - 1
                for i in range(n):
                    spaces += ' '
                answer += f'{dash}\n❗️{day}❗{spaces}|️\n{dash}\n'
                lesson = 1
                answer += (f'{lesson}📍{time}\n')
                temp = day
                lesson += 1

            if i % 2 != 0:
                answer += (f'{lesson}📍{time}\n- {even.title()}\n {sbj}\n\n')
                lesson += 1
            else:
                answer += (f'- {even.title()}\n {sbj}\n\n')

        return answer

    def get_scd_even(self, eveness = "ЧЁТ."):
        answer = f'Расписание для {self.edup} {self.course}-{self.group}\n'
        answer += f'{eveness.title()} неделя\n'
        # получаем граничные индексы
        c = self.get_indexes_cat(self.sheet, self.edup)[1]  # coll
        c = c + int(self.group) - 1                         # перезаписываем столбец на новое значение

        r = self.get_indexes(self.sheet, 'Понедельник')[0]
        temp = ''
        lesson = ''
        for i in range(r, self.sheet.max_row):

            if (str(self.sheet.cell(i, 2).value) == 'None'): break

            day = str(self.sheet.cell(i, 2).value)
            time = str(self.sheet.cell(i, 3).value)
            even = str(self.sheet.cell(i, 4).value)
            sbj = str(self.sheet.cell(i, c).value)
            if sbj == 'None': sbj = 'Занятий нет'

            dash = '--------------------------------------️'

            if temp != str(self.sheet.cell(i, 2).value):
                spaces = ''
                n = len(dash) - len('❗️' + day + '❗️') - 1
                for i in range(n):
                    spaces += ' '
                answer += f'{dash}\n❗️{day}❗{spaces}|️\n{dash}\n'
                lesson = 1
                temp = day

            if even.strip().lower() == eveness.strip().lower():
                answer += (f'{lesson}📍{time}\n{sbj}\n\n')
                lesson += 1
        return answer

    def get_scd_weekday(self, weekday = 'Понедельник'):
        answer = f'Расписание для {self.edup} {self.course}-{self.group}\n'

        spaces = ''
        dash = '--------------------------------------️'
        n = len(dash) - len('❗️' + weekday.upper() + '❗️') - 1
        for i in range(n):
            spaces += ' '
        answer += f'{dash}\n❗️{weekday.upper()}❗{spaces}|️\n{dash}\n'


        # получаем граничные индексы
        c = self.get_indexes_cat(self.sheet, self.edup)[1]  # coll
        c = c + int(self.group) - 1                         # перезаписываем столбец на новое значение
        r = self.get_indexes(self.sheet, 'Понедельник')[0]

        temp = ''
        lesson = 1
        for i in range(r, self.sheet.max_row):

            if (str(self.sheet.cell(i, 2).value) == 'None'): break

            day = str(self.sheet.cell(i, 2).value)
            time = str(self.sheet.cell(i, 3).value)
            even = str(self.sheet.cell(i, 4).value)
            sbj = str(self.sheet.cell(i, c).value)
            if sbj == 'None': sbj = 'Занятий нет'

            if weekday.lower() == day.lower():

                if i % 2 != 0:
                    answer += (f'{lesson}📍{time}\n- {even.title()}\n {sbj}\n\n')
                    lesson += 1
                else:
                    answer += (f'- {even.title()}\n {sbj}\n\n')

        return answer
